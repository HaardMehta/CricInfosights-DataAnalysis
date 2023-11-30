import requests
import openpyxl
from bs4 import BeautifulSoup

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = "Results"
print(excel.sheetnames)
sheet.append(['Team1', 'Team2', 'Winner', 'Margin', 'Ground', 'MatchDate' 'Scorecard'])

try:
    source = requests.get('https://www.espncricinfo.com/records/tournament/team-match-results/icc-men-s-t20-world-cup-2022-23-14450')
    source.raise_for_status()

    soup = BeautifulSoup(source.text, 'html.parser')
    # print(soup)

    results = soup.find('tbody', class_ = "").find_all('tr')
    
    for result in results:
        cells = result.find_all('td')
        team1 = result.find('td', class_ = "ds-min-w-max").span.text
        if len(cells) >= 7:
            team2 = cells[1].get_text()
            winner = cells[2].get_text()
            margin = cells[3].get_text()
            ground = cells[4].get_text()
            scorecard = cells[6].get_text()

        matchDate = result.find('td', class_ = "ds-min-w-max ds-text-right ds-whitespace-nowrap").span.text
        
        print(team1, team2, winner, margin, ground, matchDate, scorecard)
        sheet.append([team1, team2, winner, margin, ground, matchDate, scorecard])
        
        # break

except Exception as e:
    print(e)

excel.save("MatchResults.xlsx")